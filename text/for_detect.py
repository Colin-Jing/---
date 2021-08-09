# coding: UTF-8
import numpy as np
import torch
import pickle as pkl
from openpyxl import load_workbook
import re
import os

UNK, PAD = '<UNK>', '<PAD>'  # 未知字，padding符号

def detect(config, model, test_iter):
    result = evaluate(config, model, test_iter)
    return result


def evaluate(config, model, data_iter): # 验证集
    model.eval()
    predict_all = np.array([], dtype=int)
    with torch.no_grad():
        for texts, labels in data_iter:
            outputs = model(texts)
            predic = torch.max(outputs.data, 1)[1].cpu().numpy()
            predict_all = np.append(predict_all, predic)
    try:
        wb = load_workbook(config.test_path)
    except:
        return config.class_list[predict_all[0]]
    ws = wb[wb.sheetnames[0]]
    for i in range(1, ws.max_row):
        ws.cell(row=i+1, column=2).value = config.class_list[predict_all[i-1]]
    wb.save(config.result_path)
    return config.result_path

def build_dataset(config, ues_word):
    if ues_word:
        tokenizer = lambda x: x.split(' ')  # 以空格隔开，word-level
    else:
        tokenizer = lambda x: [y for y in x]  # char-level
    vocab = pkl.load(open(config.vocab_path, 'rb'))

    def load_dataset(path, pad_size=32):
        contents = []
        if os.path.basename(path).split('.')[-1] != 'xlsx':
            content = path.replace("THUCNews","")
            content = re.sub('[0-9’!"#$%&\'()*+,-./:：;<=>?@，；｜＃。?★、…【】？“”‘’•（）！[\\]^_`{|}~\s]+', "", content)
            content = re.sub(
                '[\001\002\003\004\005\006\007\x08\x09\x0a\x0b\x0c\x0d\x0e\x0f\x10\x11\x12\x13\x14\x15\x16\x17\x18\x19\x1a]+',
                '', content)
            label = 0
            words_line = []
            token = tokenizer(content)
            seq_len = len(token)
            if pad_size:
                if len(token) < pad_size:
                    token.extend([vocab.get(PAD)] * (pad_size - len(token)))
                else:
                    token = token[:pad_size]
                    seq_len = pad_size
            # word to id
            for word in token:
                words_line.append(vocab.get(word, vocab.get(UNK)))
            contents.append((words_line, int(label), seq_len))  # 将词变为索引
            return contents  # [([...], 0), ([...], 1), ...]

        else:
            wb = load_workbook(path)
            ws = wb[wb.sheetnames[0]]
            for i in range(2, ws.max_row + 1):
                content = ws.cell(row=i, column=3).value
                content = re.sub('[a-zA-Z0-9’!"#$%&\'()*+,-./:：;<=>?@，；｜＃。?★、…【】？“”‘’•（）！[\\]^_`{|}~\s]+', "", content)
                content = re.sub(
                    '[\001\002\003\004\005\006\007\x08\x09\x0a\x0b\x0c\x0d\x0e\x0f\x10\x11\x12\x13\x14\x15\x16\x17\x18\x19\x1a]+',
                    '', content)
                label = 0
                words_line = []
                token = tokenizer(content)
                seq_len = len(token)
                if pad_size:
                    if len(token) < pad_size:
                        token.extend([vocab.get(PAD)] * (pad_size - len(token)))
                    else:
                        token = token[:pad_size]
                        seq_len = pad_size
                # word to id
                for word in token:
                    words_line.append(vocab.get(word, vocab.get(UNK)))
                contents.append((words_line, int(label), seq_len))  # 将词变为索引
            return contents  # [([...], 0), ([...], 1), ...]
    test = load_dataset(config.test_path, config.pad_size)  # 读取测试集
    return vocab, test
