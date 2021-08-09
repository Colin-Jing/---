import pymysql
import openpyxl


class MyDataBase:
    def __init__(self, host, user, password, port, db, table):
        self.host = host
        self.user = user
        self.passwd = password
        self.port = port
        self.db = db
        self.table = table
        self.conn = None
        self.cur = None

    def connect(self):
        self.conn = pymysql.connect(host=self.host, user=self.user, passwd=self.passwd, port=self.port, db=self.db,
                                    charset='utf8')
        self.cur = self.conn.cursor()
        print("连接成功")

    def addWith(self, input, output):
        sql = "insert into %s(`input`,`output`) value(%s,%s)" % (self.table, input, output)
        self.cur.execute(sql)
        self.conn.commit()

    def addExcel(self, path):
        wb = openpyxl.load_workbook(path)
        sheet = wb['Sheet1']
        for i in range(sheet.max_row):
            tempInput = "'%s'" % sheet.cell(row=i + 1, column=1).value
            tempOutput = "'%s'" % sheet.cell(row=i + 1, column=2).value
            self.addWith(tempInput, tempOutput)

    def searchByTime(self, time):
        sql = "select * from %s where time = '%s'" % (self.table, time)
        self.cur.execute(sql)
        data = self.cur.fetchall()
        temp = [data[0][2], data[0][3]]
        print(temp)
        return temp

    def showAll(self):
        sql = "select * from %s" % self.table
        self.cur.execute(sql)
        data = self.cur.fetchall()
        for ele in data:
            print(ele)

    def close(self):
        self.cur.close()
        self.conn.close()
        print("成功关闭")


if __name__ == "__main__":
    host = "127.0.0.1"
    user = "root"
    password = "root"
    port = 3306
    db = "test"
    charset = "utf8"
    table = "output"

    mb = MyDataBase(host, user, password, port, db, table)
    mb.connect()
    mb.addWith(input="'《赛博朋克》'", output="'游戏'")
    mb.showAll()
    mb.close()
