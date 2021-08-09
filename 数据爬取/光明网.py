# 这段代码在本地运行
from bs4 import BeautifulSoup
from selenium import webdriver #从selenium库中调用webdriver模块
import time
import openpyxl
from selenium.webdriver.chrome.options import Options
import re
import random
import requests

def get_data(website,kind,url):
    time.sleep(2)
    print(url)
    try:
        res = requests.get(url) # 访问页面
        res.encoding = "utf-8"
        if res.status_code != 200:
            print(res.status_code)
            return
    except:
        print("页面不存在，该板块加载完毕！！！")
        return
    soup = BeautifulSoup(res.text, 'html.parser')
    label = soup.find_all('span',class_="channel-newsTitle")
    try:
        wb = openpyxl.load_workbook(r"./{0}/{1}.xlsx".format(website,kind))
        print("./{0}/{1}.xlsx".format(website,kind)+"被打开")
    except:
        print("./{0}/{1}.xlsx不存在！！！将自动创建".format(website,kind))
        wb = openpyxl.Workbook()
    finally:
        ws = wb.active
    # for title in label:
    #     title = title.find_all('a')
    #     for j in range(len(title)):
    #         print('{}'.format(j)+'\t'+title[j].text) # 打印label的文本
    #         ws.append([title[j].text, kind])
    for i in range(len(label)):
        if label[i].find('a',target="_blank") != None :
            if i % 5 == 0:
                print('{}'.format(i)+'\t'+label[i].find('a',target="_blank").text) # 打印label的文本
            ws.append([label[i].find('a',target="_blank").text, kind])
    wb.save(r'./{0}/{1}.xlsx'.format(website,kind))
    print(r'./{0}/{1}.xlsx'.format(website,kind)+'保存成功！！！')
    wb.close()

def get_data1(website,kind,url):
    time.sleep(2)
    print(url)
    try:
        res = requests.get(url) # 访问页面
        res.encoding = "utf-8"
        if res.status_code != 200:
            print(res.status_code)
            return
    except:
        print("页面不存在，该板块加载完毕！！！")
        return
    soup = BeautifulSoup(res.text, 'html.parser')
    label = soup.find_all('ul',class_="channel-newsGroup")
    try:
        wb = openpyxl.load_workbook(r"./{0}/{1}.xlsx".format(website,kind))
        print("./{0}/{1}.xlsx".format(website,kind)+"被打开")
    except:
        print("./{0}/{1}.xlsx不存在！！！将自动创建".format(website,kind))
        wb = openpyxl.Workbook()
    finally:
        ws = wb.active
    # for title in label:
    #     title = title.find_all('a')
    #     for j in range(len(title)):
    #         print('{}'.format(j)+'\t'+title[j].text) # 打印label的文本
    #         ws.append([title[j].text, kind])
    for i in range(len(label)):
        if label[i].find('a',target="_blank") != None :
            if i % 5 == 0:
                print('{}'.format(i)+'\t'+label[i].find('a',target="_blank").text) # 打印label的文本
            ws.append([label[i].find('a',target="_blank").text, kind])
    wb.save(r'./{0}/{1}.xlsx'.format(website,kind))
    print(r'./{0}/{1}.xlsx'.format(website,kind)+'保存成功！！！')
    wb.close()

type_list = ['教育','科技','其他']
url_list = [
    'https://edu.gmw.cn/node_10602_{}.htm',
    'https://tech.gmw.cn/node_9671_{}.htm',
    'https://life.gmw.cn/node_9268_{}.htm'
]

type_list1 = ['军事','体育']
url_list1 = [

    'https://mil.gmw.cn/node_8982_{}.htm',
    'https://sports.gmw.cn/node_9633_{}.htm',
]
'''
for i in range(len(type_list)):
    print("-*"*20)
    print('当前爬取内容：%s'%type_list[0])
    for j in range(2,11):
        if j % 5 == 0:
            print("当前爬取页面：第%d页"%j)
        get_data('光明网',type_list[0],url_list[0].format(j))
'''
for i in range(len(type_list1)):
    print("-*"*20)
    print('当前爬取内容：%s'%type_list1[i])
    for j in range(2,11):
        if j % 5 == 0:
            print("当前爬取页面：第%d页"%j)
        get_data1('光明网',type_list1[i],url_list1[i].format(j))




def hea(website,kind,url):
    chrome_options = Options()
    # 设置chrome浏览器无界面模式
    # chrome_options.add_argument('--headless')
    # chrome_options.add_argument('--disable-gpu')
    # chrome_options.add_argument('log-level=3')
    chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.128 Safari/537.36")
    driver = webdriver.Chrome(options=chrome_options)  # 设置引擎为Chrome，从本地打开一个Chrome浏览器
    driver.get(url) # 访问页面
    time.sleep(0.5)
    for i in range(760):
        driver.execute_script('window.scrollTo(0, document.body.scrollHeight)')
        time.sleep(1)
        try:
            button = driver.find_element_by_xpath('/html/body/div[5]/div[1]/ul/div/a/cite')
            button.click()
        except:
            pass
        #driver.refresh()
        if i % 10 == 0:
            print(i)
        time.sleep(1)
    HTML = driver.page_source
    bs = BeautifulSoup(HTML,'html.parser')
    label = bs.find_all('h1')  # 解析网页并提取第一个<label>标签
    try:
        wb = openpyxl.load_workbook(r"./{0}/{1}.xlsx".format(website,kind))
        print("./{0}/{1}.xlsx".format(website, kind) + "被打开")
    except:
        print("./{0}/{1}.xlsx不存在！！！将自动创建".format(website,kind))
        wb = openpyxl.Workbook()
    finally:
        ws = wb.active
    for i in range(len(label)):
        if label[i].find('a', class_='title') != None:
            if i % 10 == 0:
                print('{}'.format(i)+'\t'+label[i].find('a', class_='title').text) # 打印label的文本
            ws.append([label[i].find('a', class_='title').text, kind])
    wb.save(r'./{0}/{1}.xlsx'.format(website,kind))
    print('./{0}/{1}.xlsx'.format(website,kind)+'保存成功！！！')
    wb.close()
    driver.close() # 关闭浏览器

type_list1 = ['其他']
url_list1 = [

    'http://www.healthnet.com.cn/news.html'
]

'''
for i in range(len(type_list1)):
    print("-*"*20)
    print('当前爬取内容：%s'%type_list1[i])
    hea('光明网',type_list1[i],url_list1[i])
'''