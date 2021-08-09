# 这段代码在本地运行
from bs4 import BeautifulSoup
from selenium import webdriver #从selenium库中调用webdriver模块
import time
import openpyxl
from selenium.webdriver.chrome.options import Options

def get_data(website,kind,url):
    chrome_options = Options()
    # 设置chrome浏览器无界面模式
    chrome_options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.128 Safari/537.36")
    driver = webdriver.Chrome(options=chrome_options)  # 设置引擎为Chrome，从本地打开一个Chrome浏览器
    driver.get(url) # 访问页面
    time.sleep(0.5)
    for i in range(2,200):
        for j in [1,2]:
            driver.execute_script('window.scrollTo(0, document.body.scrollHeight)')
        time.sleep(1.5)
        button = driver.find_element_by_xpath('/html/body/div[9]/div[2]/div[1]/div[3]/span/a[9]')
        #ActionChains(self.driver).move_to_element(element).perform()
        try:
            button.click()
        except:
            pass
        #driver.refresh()
        if i % 10 == 0:
            print(i)
        time.sleep(1.5)
        HTML = driver.page_source
        bs = BeautifulSoup(HTML,'html.parser')
        label = bs.find_all('div',class_="tit")  # 解析网页并提取第一个<label>标签
        try:
            wb = openpyxl.load_workbook(r"./{0}/{1}.xlsx".format(website,kind))
            print("./{0}/{1}.xlsx".format(website, kind) + "被打开")
        except:
            print("./{0}/{1}.xlsx不存在！！！将自动创建".format(website,kind))
            wb = openpyxl.Workbook()
        finally:
            ws = wb.active
        for i in range(len(label)):
            if label[i].find('a', target="_blank") != None:
                if i % 10 == 0:
                    print('{}'.format(i)+'\t'+label[i].find('a',target="_blank").text) # 打印label的文本
                ws.append([label[i].find('a',target="_blank").text, kind])
        wb.save(r'./{0}/{1}.xlsx'.format(website,kind))
        print('./{0}/{1}.xlsx'.format(website,kind)+'保存成功！！！')
        wb.close()
    driver.close() # 关闭浏览器

def get_data1(website,kind,url):
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.128 Safari/537.36")
    driver = webdriver.Chrome(options=chrome_options) # 设置引擎为Chrome，从本地打开一个Chrome浏览器
    driver.get(url) # 访问页面
    time.sleep(0.5)
    for i in range(2,200):
        for j in [1,2]:
            driver.execute_script('window.scrollTo(0, document.body.scrollHeight)')
        time.sleep(1.5)
        button = driver.find_element_by_xpath('/html/body/div[7]/div[2]/div[1]/div[3]/span/a[9]')
        #ActionChains(self.driver).move_to_element(element).perform()
        try:
            button.click()
        except:
            pass
        #driver.refresh()
        if i % 10 == 0:
            print(i)
        time.sleep(1.5)
        HTML = driver.page_source
        bs = BeautifulSoup(HTML,'html.parser')
        label = bs.find_all('div',class_="tit")  # 解析网页并提取第一个<label>标签
        try:
            wb = openpyxl.load_workbook(r"./{0}/{1}.xlsx".format(website,kind))
            print("./{0}/{1}.xlsx".format(website, kind) + "被打开")
        except:
            print("./{0}/{1}.xlsx不存在！！！将自动创建".format(website,kind))
            wb = openpyxl.Workbook()
        finally:
            ws = wb.active
        for i in range(len(label)):
            if label[i].find('a', target="_blank") != None:
                if i % 10 == 0:
                    print('{}'.format(i)+'\t'+label[i].find('a',target="_blank").text) # 打印label的文本
                ws.append([label[i].find('a',target="_blank").text, kind])
        wb.save(r'./{0}/{1}.xlsx'.format(website,kind))
        print('./{0}/{1}.xlsx'.format(website,kind)+'保存成功！！！')
        wb.close()
    driver.close() # 关闭浏览器


type_list = ['游戏']
url_list = [
   'https://www.gamersky.com/news/pc/zx/'
]

type_list1 = ['游戏']
url_list1 = [

    'https://ol.gamersky.com/news/game/'
]

# for i in range(len(type_list)):
#     print("-*"*20)
#     print('当前爬取内容：%s'%type_list[i])
#     get_data('游民星空',type_list[i],url_list[i])


for i in range(len(type_list1)):
    print("-*"*20)
    print('当前爬取内容：%s'%type_list1[i])
    get_data1('游民星空',type_list1[i],url_list1[i])