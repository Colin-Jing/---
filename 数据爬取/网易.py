# 这段代码在本地运行
from bs4 import BeautifulSoup
import time
import openpyxl
import requests


def get_data(website,kind,url):
    time.sleep(2)
    print(url)
    try:
        res = requests.get(url) # 访问页面
        res.encoding = "gbk"
        if res.status_code != 200:
            print(res.status_code)
            return
    except:
        print("页面不存在，该板块加载完毕！！！")
        return
    soup = BeautifulSoup(res.text, 'html.parser')
    label = soup.find_all('div',class_="titleBar clearfix")
    try:
        wb = openpyxl.load_workbook(r"./{0}/{1}.xlsx".format(website,kind))
        print("./{0}/{1}.xlsx".format(website,kind)+"被打开")
    except:
        print("./{0}/{1}.xlsx不存在！！！将自动创建".format(website,kind))
        wb = openpyxl.Workbook()
    finally:
        ws = wb.active
    for i in range(len(label)):
        if label[i].find('a') != None :
            if i % 5 == 0:
                print('{}'.format(i)+'\t'+label[i].find('a').text) # 打印label的文本
            ws.append([label[i].find('a').text, kind])
    wb.save(r'./{0}/{1}.xlsx'.format(website,kind))
    print(r'./{0}/{1}.xlsx'.format(website,kind)+'保存成功！！！')
    wb.close()

def get_data1(website,kind,url):
    time.sleep(2)
    print(url)
    try:
        res = requests.get(url) # 访问页面
        res.encoding = "gbk"
        if res.status_code != 200:
            print(res.status_code)
            return
    except:
        print("页面不存在，该板块加载完毕！！！")
        return
    soup = BeautifulSoup(res.text, 'html.parser')
    label = soup.find_all('div',class_="item-cont")
    try:
        wb = openpyxl.load_workbook(r"./{0}/{1}.xlsx".format(website,kind))
        print("./{0}/{1}.xlsx".format(website,kind)+"被打开")
    except:
        print("./{0}/{1}.xlsx不存在！！！将自动创建".format(website,kind))
        wb = openpyxl.Workbook()
    finally:
        ws = wb.active
    for i in range(len(label)):
        if label[i].find('a') != None :
            if i % 5 == 0:
                print('{}'.format(i)+'\t'+label[i].find('a').text) # 打印label的文本
            ws.append([label[i].find('a').text, kind])
    wb.save(r'./{0}/{1}.xlsx'.format(website,kind))
    print(r'./{0}/{1}.xlsx'.format(website,kind)+'保存成功！！！')
    wb.close()

def get_data2(website,kind,url):
    time.sleep(2)
    print(url)
    try:
        res = requests.get(url) # 访问页面
        res.encoding = "gbk"
        if res.status_code != 200:
            print(res.status_code)
            return
    except:
        print("页面不存在，该板块加载完毕！！！")
        return
    soup = BeautifulSoup(res.text, 'html.parser')
    label = soup.find_all('ul',class_="articleList")
    try:
        wb = openpyxl.load_workbook(r"./{0}/{1}.xlsx".format(website,kind))
        print("./{0}/{1}.xlsx".format(website,kind)+"被打开")
    except:
        print("./{0}/{1}.xlsx不存在！！！将自动创建".format(website,kind))
        wb = openpyxl.Workbook()
    finally:
        ws = wb.active
    for title in label:
        title = title.find_all('a')
        for j in range(len(title)):
            print('{}'.format(j)+'\t'+title[j].text) # 打印label的文本
            ws.append([title[j].text, kind])

    wb.save(r'./{0}/{1}.xlsx'.format(website,kind))
    print(r'./{0}/{1}.xlsx'.format(website,kind)+'保存成功！！！')
    wb.close()

type_list = ['科技']
url_list = [
    'https://tech.163.com/special/gd2016_{}/'
]

type_list1 = ['汽车']
url_list1 = [
    'https://auto.163.com/special/2016news_{}/#subtab'
]

type_list2 = ['体育']
url_list2 = [
    'http://sports.163.com/special/0005rt/sportsgd_{}.html'
]
#
# for i in range(len(type_list)):
#     print("-*"*20)
#     print('当前爬取内容：%s'%type_list[i])
#     for j in range(2,10):
#         if j % 5 == 0:
#             print("当前爬取页面：第%d页"%j)
#         get_data('网易',type_list[i],url_list[i].format('0'+str(j)))
#     for j in range(10,21):
#         if j % 5 == 0:
#             print("当前爬取页面：第%d页"%j)
#         get_data('网易',type_list[i],url_list[i].format(j))
'''
for i in range(len(type_list1)):
    print("-*"*20)
    print('当前爬取内容：%s'%type_list1[i])
    for j in range(1,10):
        if j % 5 == 0:
            print("当前爬取页面：第%d页"%j)
        get_data1('网易',type_list1[i],url_list1[i].format('0'+str(j)))
    for j in range(10,16):
        if j % 5 == 0:
            print("当前爬取页面：第%d页"%j)
        get_data1('网易',type_list1[i],url_list1[i].format(j))
'''
for i in range(len(type_list2)):
    print("-*"*20)
    print('当前爬取内容：%s'%type_list2[i])
    for j in range(1,10):
        if j % 5 == 0:
            print("当前爬取页面：第%d页"%j)
        get_data2('网易',type_list2[i],url_list2[i].format('0'+str(j)))
    for j in range(10,16):
        if j % 5 == 0:
            print("当前爬取页面：第%d页"%j)
        get_data2('网易',type_list2[i],url_list2[i].format(j))