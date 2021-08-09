from bs4 import BeautifulSoup
import time
import openpyxl
import requests
from selenium.webdriver.chrome.options import Options
from selenium import webdriver #从selenium库中调用webdriver模块


#教育类爬虫
def edu(website,kind,url):
    time.sleep(2)
    print(url)
    try:
        res = requests.get(url) # 访问页面
        res.encoding = "UTF-8"
        if res.status_code != 200:
            print(res.status_code)
            return
    except:
        print("页面不存在，该板块加载完毕！！！")
        return
    soup = BeautifulSoup(res.text, 'html.parser')
    label = soup.find_all('article',class_='excerpt excerpt-1')
    try:
        wb = openpyxl.load_workbook(r"./{0}/{1}.xlsx".format(website,kind))
        print("./{0}/{1}.xlsx".format(website,kind)+"被打开")
    except:
        print("./{0}/{1}.xlsx不存在！！！将自动创建".format(website,kind))
        wb = openpyxl.Workbook()
    finally:
        ws = wb.active
    # for title in label:
    #     title = title.find_all('a')
    #     for j in range(len(title)):
    #         print('{}'.format(j)+'\t'+title[j].text) # 打印label的文本
    #         ws.append([title[j].text, kind])
    for i in range(len(label)):
        if label[i].find('a',target="_blank")['title'] != None :
            if i % 5 == 0:
                print('{}'.format(i)+'\t'+label[i].find('a',target="_blank")['title']) # 打印label的文本
            ws.append([label[i].find('a',target="_blank")['title'], kind])
    wb.save(r'./{0}/{1}.xlsx'.format(website,kind))
    print(r'./{0}/{1}.xlsx'.format(website,kind)+'保存成功！！！')
    wb.close()

type_list = ['教育']
url_list = [
    'https://www.edunews.net.cn/news/{}.html',
]
for i in range(len(type_list)):
    print("-*"*20)
    print('当前爬取内容：%s'%type_list[i])
    for j in range(2,54):
        if j % 5 == 0:
            print("当前爬取页面：第%d页"%j)
        edu('人民网',type_list[i],url_list[i].format(j))


#科技类爬虫




#军事类爬虫
def get_data(kind,url):
    print(url)
    chrome_options = Options()
    # 设置chrome浏览器无界面模式
    chrome_options.add_argument('--headless')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('log-level=3')
    # chrome_options.add_argument('user-agent="Mozilla/5.0 (iPod; U; CPU iPhone OS 2_1 \
    #                             like Mac OS X; ja-jp) AppleWebKit/525.18.1 (KHTML, like \
    #                             Gecko) Version/3.1.1 Mobile/5F137 Safari/525.20"')
    driver = webdriver.Chrome(options=chrome_options) # 设置引擎为Chrome，从本地打开一个Chrome浏览器
    driver.get(url) # 访问页面
    print('页面加载完毕！！！')
    for i in range(2000):
        driver.execute_script('window.scrollTo(0, document.body.scrollHeight)')
        if i % 10 == 0:
            print(i)
        time.sleep(1)
    HTML = driver.page_source
    bs = BeautifulSoup(HTML,'html.parser')
    print(bs.text)
    label = bs.find_all('div',class_="con-txt")  # 解析网页并提取第一个<label>标签

    try:
        wb = openpyxl.load_workbook(r"./腾讯网/{}.xlsx".format(kind))
        print("./腾讯网/{}.xlsx".format(kind) + "被打开")
    except:
        print("./腾讯网/{}.xlsx不存在！！！将自动创建".format(kind))
        wb = openpyxl.Workbook()
    finally:
        ws = wb.active
    for i in range(len(label)):
        if i % 10 == 0:
            print('{}'.format(i)+'\t'+label[i].find('h4').text) # 打印label的文本
        ws.append([label[i].find('h4').text, kind])
    wb.save(r'./腾讯网/{}.xlsx'.format(kind))
    print('./腾讯网/{}.xlsx'.format(kind) + '保存成功！！！')
    wb.close()
    driver.close() # 关闭浏览器


type_list = ['军事']
url_list = [
    'https://mil.huanqiu.com/',
]
for i in range(len(type_list)):
    print("-*"*20)
    print('当前爬取内容：%s'%type_list[i])
    get_data(type_list[i],url_list[i])



#汽车类爬虫


#体育类爬虫
def sports(website,kind,url):
    chrome_options = Options()
    #设置chrome浏览器无界面模式
    chrome_options.add_argument('--headless')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('log-level=3')
    chrome_options.add_argument('user-agent="Mozilla/5.0 (iPod; U; CPU iPhone OS 2_1 \
                                like Mac OS X; ja-jp) AppleWebKit/525.18.1 (KHTML, like \
                                Gecko) Version/3.1.1 Mobile/5F137 Safari/525.20"')
    driver = webdriver.Chrome(options=chrome_options) # 设置引擎为Chrome，从本地打开一个Chrome浏览器
    try:
        driver.get(url) # 访问页面
        print(url)
    except:
        print('访问失败')
        return

    driver.execute_script('window.scrollTo(0, document.body.scrollHeight)')
    HTML = driver.page_source
    bs = BeautifulSoup(HTML,'html.parser')
    label = bs.find_all('li',class_="ty-card-type10-makeup")  # 解析网页并提取第一个<label>标签
    try:
        wb = openpyxl.load_workbook(r"./{0}/{1}.xlsx".format(website,kind))
        print("./{0}/{1}.xlsx".format(website, kind) + "被打开")
    except:
        print("./{0}/{1}.xlsx不存在！！！将自动创建".format(website,kind))
        wb = openpyxl.Workbook()
    finally:
        ws = wb.active
    for i in range(len(label)):
        if label[i].find('a', target="_blank") != None:
            if i % 10 == 0:
                print('{}'.format(i)+'\t'+label[i].find('a',target="_blank").text) # 打印label的文本
            ws.append([label[i].find('a',target="_blank").text, kind])
    wb.save(r'./{0}/{1}.xlsx'.format(website,kind))
    print('./{0}/{1}.xlsx'.format(website,kind)+'保存成功！！！')
    wb.close()
    driver.close() # 关闭浏览器


type_list = ['体育','体育']
url_list = [
    'https://sports.sina.com.cn/head/sports2020{0}{1}am.shtml',
    'https://sports.sina.com.cn/head/sports2020{0}{1}pm.shtml'
]
'''

for i in range(len(type_list)):
    print("-*"*20)
    print('当前爬取内容：%s'%type_list[i])
    for j in range(7,13):
        for m in range(1,10):
            print(str(j)+'月'+str(m)+'号')
            sports('新浪网',type_list[i],url_list[i].format('0'+str(j),'0'+str(m)))
        for n in range(10,29):
            print(str(j) + '月' + str(n) + '号')
            sports('新浪网', type_list[i], url_list[i].format('0'+str(j),str(n)))
'''

#娱乐类爬虫

def ent(website,kind,url):
    chrome_options = Options()
    #设置chrome浏览器无界面模式
    chrome_options.add_argument('--headless')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('log-level=3')
    # chrome_options.add_argument('user-agent="Mozilla/5.0 (iPod; U; CPU iPhone OS 2_1 \
    #                             like Mac OS X; ja-jp) AppleWebKit/525.18.1 (KHTML, like \
    #                             Gecko) Version/3.1.1 Mobile/5F137 Safari/525.20"')
    driver = webdriver.Chrome(options=chrome_options) # 设置引擎为Chrome，从本地打开一个Chrome浏览器
    try:
        driver.get(url) # 访问页面
        print(url)
    except:
        print('访问失败')
        return
    time.sleep(0.3)
    driver.execute_script('window.scrollTo(0, document.body.scrollHeight)')
    time.sleep(0.5)
    # button = driver.find_element_by_xpath('/html/body/div[9]/div[2]/div[1]/div[3]/span/a[9]')
    #
    # try:
    #     button.click()
    # except:
    #     pass
    HTML = driver.page_source
    bs = BeautifulSoup(HTML,'html.parser')
    label = bs.find_all('td',class_="ConsTi")  # 解析网页并提取第一个<label>标签
    try:
        wb = openpyxl.load_workbook(r"./{0}/{1}.xlsx".format(website,kind))
        print("./{0}/{1}.xlsx".format(website, kind) + "被打开")
    except:
        print("./{0}/{1}.xlsx不存在！！！将自动创建".format(website,kind))
        wb = openpyxl.Workbook()
    finally:
        ws = wb.active
    for i in range(len(label)):
        if label[i].find('a', target="_blank") != None:
            if i % 10 == 0:
                print('{}'.format(i)+'\t'+label[i].find('a',target="_blank").text) # 打印label的文本
            ws.append([label[i].find('a',target="_blank").text, kind])
    wb.save(r'./{0}/{1}.xlsx'.format(website,kind))
    print('./{0}/{1}.xlsx'.format(website,kind)+'保存成功！！！')
    wb.close()
    driver.close() # 关闭浏览器



type_list = ['娱乐']
url_list = [
    'http://ent.sina.com.cn/topnews/2021{0}{1}.html',
]
# for i in range(len(type_list)):
#     print("-*"*20)
#     print('当前爬取内容：%s'%type_list[i])
#     for j in range(1):
#         if j % 5 == 0:
#             print("当前爬取页面：第%d页"%j)
#         ent('新浪网',type_list[i],url_list[i])

# for i in range(len(type_list)):
#     print("-*"*20)
#     print('当前爬取内容：%s'%type_list[i])
#     for j in range(1,7):
#         for m in range(1,10):
#             print(str(j)+'月'+str(m)+'号')
#             ent('新浪网',type_list[i],url_list[i].format('0'+str(j),'0'+str(m)))
#         for n in range(10,29):
#             print(str(j) + '月' + str(n) + '号')
#             ent('新浪网', type_list[i], url_list[i].format('0'+str(j),str(n)))

#其他类爬虫



















