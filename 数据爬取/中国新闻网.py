# 这段代码在本地运行
from bs4 import BeautifulSoup
from selenium import webdriver #从selenium库中调用webdriver模块
import time
import openpyxl
import requests


def get_data(website,kind,url):
    driver = webdriver.Chrome() # 设置引擎为Chrome，从本地打开一个Chrome浏览器
    driver.get(url) # 访问页面
    time.sleep(0.5)
    for i in range(300):
        driver.execute_script('window.scrollTo(0, document.body.scrollHeight)')
        try:
            button = driver.find_element_by_id('page_bar0')
            button.click()
        except:
            print('加载完毕！！！')
            break
        if i % 10 == 0:
            print(i)
        time.sleep(1.5)
    HTML = driver.page_source
    time.sleep(1) # 等待3秒
    bs = BeautifulSoup(HTML,'html.parser')
    label = bs.find_all('div',class_="news_title")  # 解析网页并提取第一个<label>标签
    try:
        wb = openpyxl.load_workbook(r"./{0}/{1}.xlsx".format(website,kind))
    except:
        print("./{0}/{1}.xlsx不存在！！！将自动创建".format(website,kind))
        wb = openpyxl.Workbook()
    finally:
        ws = wb.active
    for i in range(len(label)):
        if i % 10 == 0:
            print('{}'.format(i)+'\t'+label[i].find('a').text) # 打印label的文本
        ws.append([label[i].find('a').text, kind])
    wb.save(r'./{0}/{1}.xlsx'.format(website,kind))
    wb.close()
    driver.close() # 关闭浏览器


type_list = ['财经','汽车','体育','其他']
url_list = [
    'https://www.chinanews.com/finance/',
    'https://www.chinanews.com/auto/',
    'https://www.chinanews.com/sports/',
    'https://www.chinanews.com/life/'
]
'''
for i in range(len(type_list)):
    print("-*"*20)
    print('当前爬取内容：%s'%type_list[0])
    get_data('中国新闻网',type_list[0],url_list[0])
'''
def get_data2(website,kind,url):
    time.sleep(0.3)
    print(url)
    headers = {
        'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/536.5 (KHTML, like Gecko) Chrome/19.0.1084.9 Safari/536.5'
    }

    try:
        res = requests.get(url,headers=headers) # 访问页面
        res.encoding = "GB2312"
        if res.status_code != 200:
            print(res.status_code)
            return
    except:
        print("页面不存在，该板块加载完毕！！！")
        return
    soup = BeautifulSoup(res.text, 'html.parser')
    label = soup.find_all('li')
    try:
        wb = openpyxl.load_workbook(r"./{0}/{1}.xlsx".format(website,kind))
        print("./{0}/{1}.xlsx".format(website,kind)+"被打开")
    except:
        print("./{0}/{1}.xlsx不存在！！！将自动创建".format(website,kind))
        wb = openpyxl.Workbook()
    finally:
        ws = wb.active
    for i in range(len(label)):
        if label[i].find('a', target="_blank") != None:
            if i % 10 == 0:
                print('{}'.format(i) + '\t' + label[i].find('a', target="_blank").text)  # 打印label的文本
            ws.append([label[i].find('a', target="_blank").text, kind])
    wb.save(r'./{0}/{1}.xlsx'.format(website, kind))
    print('./{0}/{1}.xlsx'.format(website, kind) + '保存成功！！！')
    wb.close()



def get_data3(website,kind,url):
    time.sleep(0.3)
    print(url)
    headers = {
        'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/536.5 (KHTML, like Gecko) Chrome/19.0.1084.9 Safari/536.5'
    }

    try:
        res = requests.get(url,headers=headers) # 访问页面
        res.encoding = "gb2312"
        if res.status_code != 200:
            print(res.status_code)
            return
    except:
        print("页面不存在，该板块加载完毕！！！")
        return
    soup = BeautifulSoup(res.text, 'html.parser')
    label = soup.find_all('span',class_='text kuaixun')
    try:
        wb = openpyxl.load_workbook(r"./{0}/{1}.xlsx".format(website,kind))
        print("./{0}/{1}.xlsx".format(website,kind)+"被打开")
    except:
        print("./{0}/{1}.xlsx不存在！！！将自动创建".format(website,kind))
        wb = openpyxl.Workbook()
    finally:
        ws = wb.active
    for i in range(len(label)):
        if label[i].find('strong') != None:
            if i % 10 == 0:
                print('{}'.format(i) + '\t' + label[i].find('strong').text)  # 打印label的文本
            ws.append([label[i].find('strong').text, kind])
    wb.save(r'./{0}/{1}.xlsx'.format(website, kind))
    print('./{0}/{1}.xlsx'.format(website, kind) + '保存成功！！！')
    wb.close()

type_list = ['科技']
url_list = [
    'http://scitech.people.com.cn/GB/1057/index{}.html',
    'http://tech.cnr.cn/techyw/technews/index{}.html'
]
type_list1 = ['科技']
url_list1 = [
    'http://tech.cnr.cn/techyw/technews/index_{}.html'
]
'''
for i in range(len(type_list)):
    print("-*"*20)
    print('当前爬取内容：%s'%type_list[0])
    for j in range(1,18):
        if j % 5 == 0:
            print("当前爬取页面：第%d页"%j)
        get_data2('央广网',type_list[0],url_list[0].format(j))
'''
for i in range(len(type_list1)):
    print("-*"*20)
    print('当前爬取内容：%s'%type_list1[0])
    for j in range(1,1000):
        if j % 5 == 0:
            print("当前爬取页面：第%d页"%j)
        get_data3('央广网',type_list1[0],url_list1[0].format(j))