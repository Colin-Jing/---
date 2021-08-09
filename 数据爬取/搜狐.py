# 这段代码在本地运行
from bs4 import BeautifulSoup
import time
import openpyxl
import requests



def get_data1(website,kind,url):
    time.sleep(0.5)
    print(url)
    try:
        res = requests.get(url) # 访问页面
        res.encoding = "utf-8"
        if res.status_code != 200:
            print(res.status_code)
            return
    except:
        print("页面不存在，该板块加载完毕！！！")
        return
    soup = BeautifulSoup(res.text, 'html.parser')
    label = soup.find_all('div',class_="news-list-detail")
    try:
        wb = openpyxl.load_workbook(r"./{0}/{1}.xlsx".format(website,kind))
        print("./{0}/{1}.xlsx".format(website,kind)+"被打开")
    except:
        print("./{0}/{1}.xlsx不存在！！！将自动创建".format(website,kind))
        wb = openpyxl.Workbook()
    finally:
        ws = wb.active
    # for title in label:
    #     title = title.find_all('a')
    #     for j in range(len(title)):
    #         print('{}'.format(j)+'\t'+title[j].text) # 打印label的文本
    #         ws.append([title[j].text, kind])
    for i in range(len(label)):
        if label[i].find('a',target="_blank") != None :
            if i % 5 == 0:
                print('{}'.format(i)+'\t'+label[i].find('a',target="_blank").text) # 打印label的文本
            ws.append([label[i].find('a',target="_blank").text, kind])
    wb.save(r'./{0}/{1}.xlsx'.format(website,kind))
    print(r'./{0}/{1}.xlsx'.format(website,kind)+'保存成功！！！')
    wb.close()

type_list = ['房产']
url_list = [
    'https://zixun.focus.cn/{}/'
]

type_list1 = ['财经']
url_list1 = [

    'https://zixun.focus.cn/caijing/{}/',
]

# for i in range(len(type_list)):
#     print("-*"*20)
#     print('当前爬取内容：%s'%type_list[i])
#     for j in range(1,1000):
#         if j % 5 == 0:
#             print("当前爬取页面：第%d页"%j)
#         get_data1('搜狐',type_list[i],url_list[i].format(j))
#
# for i in range(len(type_list1)):
#     print("-*"*20)
#     print('当前爬取内容：%s'%type_list1[i])
#     for j in range(1,462):
#         if j % 5 == 0:
#             print("当前爬取页面：第%d页"%j)
#         get_data1('搜狐',type_list1[i],url_list1[i].format(j))

# 这段代码在本地运行
from bs4 import BeautifulSoup
from selenium import webdriver #从selenium库中调用webdriver模块
import time
import openpyxl
from selenium.webdriver.chrome.options import Options

def get_data(kind,url):
    chrome_options = Options()
    # 设置chrome浏览器无界面模式
    chrome_options.add_argument('--headless')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('log-level=3')
    driver = webdriver.Chrome(options=chrome_options) # 设置引擎为Chrome，从本地打开一个Chrome浏览器
    driver.get(url) # 访问页面
    time.sleep(3)
    for i in range(1500):
        driver.execute_script('window.scrollTo(0, document.body.scrollHeight)')
        if i % 10 == 0:
            print(i)
        time.sleep(1.5)
    HTML = driver.page_source
    time.sleep(1) # 等待3秒
    bs = BeautifulSoup(HTML,'html.parser')
    label = bs.find_all('div',class_="info_area")  # 解析网页并提取第一个<label>标签
    try:
        wb = openpyxl.load_workbook(r"./搜狐/{}.xlsx".format(kind))
        print("./搜狐/{}.xlsx".format(kind) + "被打开")
    except:
        print("./搜狐/{}.xlsx不存在！！！将自动创建".format(kind))
        wb = openpyxl.Workbook()
    finally:
        ws = wb.active
    for i in range(len(label)):
        if label[i].find('h3') != None:
            if i % 10 == 0:
                print('{}'.format(i)+'\t'+label[i].find('h3').text) # 打印label的文本
            ws.append([label[i].find('h3').text, kind])
    wb.save(r'./搜狐/{}.xlsx'.format(kind))
    print('./搜狐/{}.xlsx'.format(kind) + '保存成功！！！')
    wb.close()
    driver.close() # 关闭浏览器


type_list = ['汽车']
url_list = [
    'http://auto.sohu.com/qichexinwen.shtml',
]
for i in range(len(type_list)):
    print("-*"*20)
    print('当前爬取内容：%s'%type_list[i])
    get_data(type_list[i],url_list[i])


type_list1 = ['汽车']
url_list1 = [
    'http://roll.sohu.com/auto/index{}.shtml',
]
for i in range(len(type_list1)):
    print("-*"*20)
    print('当前爬取内容：%s'%type_list1[i])
    get_data(type_list1[i],url_list1[i])