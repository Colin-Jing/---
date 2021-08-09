# 这段代码在本地运行
from bs4 import BeautifulSoup
from selenium import webdriver #从selenium库中调用webdriver模块
import time
import openpyxl
from selenium.webdriver.chrome.options import Options
import re
import random
import requests

def get_data(website,kind,url):
    time.sleep(2)
    print(url)
    proxy_list = [
        '175.146.215.206:9999',
        '118.212.106.203:9999',
        '113.237.5.111:9999',
        '27.40.91.79:9999',
        '113.194.134.86:9999'
    ]

    # 收集到的常用Header
    my_headers = [
        "Mozilla/5.0 (Windows NT 6.3; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_2) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/35.0.1916.153 Safari/537.36",
        "Mozilla/5.0 (Windows NT 6.1; WOW64; rv:30.0) Gecko/20100101 Firefox/30.0",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_2) AppleWebKit/537.75.14 (KHTML, like Gecko) Version/7.0.3 Safari/537.75.14",
        "Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Win64; x64; Trident/6.0)",
        'Mozilla/5.0 (Windows; U; Windows NT 5.1; it; rv:1.8.1.11) Gecko/20071127 Firefox/2.0.0.11',
        'Opera/9.25 (Windows NT 5.1; U; en)',
        'Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1; SV1; .NET CLR 1.1.4322; .NET CLR 2.0.50727)',
        'Mozilla/5.0 (compatible; Konqueror/3.5; Linux) KHTML/3.5.5 (like Gecko) (Kubuntu)',
        'Mozilla/5.0 (X11; U; Linux i686; en-US; rv:1.8.0.12) Gecko/20070731 Ubuntu/dapper-security Firefox/1.5.0.12',
        'Lynx/2.8.5rel.1 libwww-FM/2.14 SSL-MM/1.4.1 GNUTLS/1.2.9',
        "Mozilla/5.0 (X11; Linux i686) AppleWebKit/535.7 (KHTML, like Gecko) Ubuntu/11.04 Chromium/16.0.912.77 Chrome/16.0.912.77 Safari/535.7",
        "Mozilla/5.0 (X11; Ubuntu; Linux i686; rv:10.0) Gecko/20100101 Firefox/10.0 "
    ]
    proxy = random.choice(proxy_list)
    header = random.choice(my_headers)
    proxies = {
        'https':proxy
    }
    headers = {
    'User - Agent':header
    }

    try:
        res = requests.get(url,headers=headers,proxies=proxies) # 访问页面
        res.encoding = "gb2312"
        if res.status_code != 200:
            print(res.status_code)
            return
    except:
        print("页面不存在，该板块加载完毕！！！")
        return
    soup = BeautifulSoup(res.text, 'html.parser')
    label = soup.find_all('span',class_="text kuaixun")
    try:
        wb = openpyxl.load_workbook(r"./{0}/{1}.xlsx".format(website,kind))
        print("./{0}/{1}.xlsx".format(website,kind)+"被打开")
    except:
        print("./{0}/{1}.xlsx不存在！！！将自动创建".format(website,kind))
        wb = openpyxl.Workbook()
    finally:
        ws = wb.active
    # for title in label:
    #     title = title.find_all('a')
    #     for j in range(len(title)):
    #         print('{}'.format(j)+'\t'+title[j].text) # 打印label的文本
    #         ws.append([title[j].text, kind])
    for i in range(len(label)):
        if label[i].find('strong') != None :
            if i % 5 == 0:
                print('{}'.format(i)+'\t'+label[i].find('strong').text) # 打印label的文本
            ws.append([label[i].find('strong').text, kind])
    wb.save(r'./{0}/{1}.xlsx'.format(website,kind))
    print(r'./{0}/{1}.xlsx'.format(website,kind)+'保存成功！！！')
    wb.close()

def get_data1(website,kind,url):
    time.sleep(2)
    print(url)
    try:
        res = requests.get(url) # 访问页面
        res.encoding = "gb2312"
        if res.status_code != 200:
            print(res.status_code)
            return
    except:
        print("页面不存在，该板块加载完毕！！！")
        return
    soup = BeautifulSoup(res.text, 'html.parser')
    label = soup.find_all('span',class_="bt")
    try:
        wb = openpyxl.load_workbook(r"./{0}/{1}.xlsx".format(website,kind))
        print("./{0}/{1}.xlsx".format(website,kind)+"被打开")
    except:
        print("./{0}/{1}.xlsx不存在！！！将自动创建".format(website,kind))
        wb = openpyxl.Workbook()
    finally:
        ws = wb.active
    # for title in label:
    #     title = title.find_all('a')
    #     for j in range(len(title)):
    #         print('{}'.format(j)+'\t'+title[j].text) # 打印label的文本
    #         ws.append([title[j].text, kind])
    for i in range(len(label)):
        if label[i].find('a',target="_blank") != None :
            if i % 5 == 0:
                print('{}'.format(i)+'\t'+label[i].find('a',target="_blank").text) # 打印label的文本
            ws.append([label[i].find('a',target="_blank").text, kind])
    wb.save(r'./{0}/{1}.xlsx'.format(website,kind))
    print(r'./{0}/{1}.xlsx'.format(website,kind)+'保存成功！！！')
    wb.close()

def get_data2(website,kind,url):
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.128 Safari/537.36")
    driver = webdriver.Chrome(options=chrome_options) # 设置引擎为Chrome，从本地打开一个Chrome浏览器
    driver.get(url) # 访问页面
    time.sleep(0.5)
    for i in range(200):
        driver.execute_script('window.scrollTo(0, document.body.scrollHeight)')
        time.sleep(0.3)
        try:
            if i > 0:
                button = driver.find_element_by_xpath('/html/body/div[5]/div[4]/div[2]/div[1]/div[2]/div/div[3]/div[1]/table/tbody/tr/td/div/ul/li[12]/a')
                button.click()
            else:
                pass
        except:
            pass
        #driver.refresh()
        if i % 10 == 0:
            print(i)
        time.sleep(1.5)
        HTML = driver.page_source
        bs = BeautifulSoup(HTML,'html.parser')
        label = bs.find_all('li')  # 解析网页并提取第一个<label>标签
        try:
            wb = openpyxl.load_workbook(r"./{0}/{1}.xlsx".format(website,kind))
            print("./{0}/{1}.xlsx".format(website, kind) + "被打开")
        except:
            print("./{0}/{1}.xlsx不存在！！！将自动创建".format(website,kind))
            wb = openpyxl.Workbook()
        finally:
            ws = wb.active
        for i in range(len(label)):
            if label[i].find('a', target="_blank") != None:
                if i % 10 == 0:
                    print('{}'.format(i)+'\t'+label[i].find('a',target="_blank").text) # 打印label的文本
                ws.append([label[i].find('a',target="_blank").text, kind])
        wb.save(r'./{0}/{1}.xlsx'.format(website,kind))
        print('./{0}/{1}.xlsx'.format(website,kind)+'保存成功！！！')
        wb.close()
    driver.close() # 关闭浏览器

type_list = ['财经','教育','科技','其他']
url_list = [
    'http://finance.cnr.cn/2014jingji/yw/index_{}.html',
    'http://edu.cnr.cn/list/index_{}.html',
    'http://tech.cnr.cn/techyw/technews/index_{}.html',
    'http://health.cnr.cn/jkgdxw/index_{}.html'
]

type_list1 = ['军事']
url_list1 = [

    'http://military.cnr.cn/gz/index_{}.html',
]

for i in range(len(type_list)):
    print("-*"*20)
    print('当前爬取内容：%s'%type_list[0])
    for j in range(1,100):
        if j % 5 == 0:
            print("当前爬取页面：第%d页"%j)
        get_data('央广网',type_list[0],url_list[0].format(j))

# for i in range(len(type_list1)):
#     print("-*"*20)
#     print('当前爬取内容：%s'%type_list1[i])
#     for j in range(1,19):
#         if j % 5 == 0:
#             print("当前爬取页面：第%d页"%j)
#         get_data1('央广网',type_list1[i],url_list1[i].format(j))

# type_list1 = ['科技']
# url_list1 = [
#     'http://app.tech.china.com.cn/news/live.php?channel=%E7%A7%91%E6%8A%80&p={}',
# ]
#
# for i in range(len(type_list1)):
#     print("-*"*20)
#     print('当前爬取内容：%s'%type_list1[i])
#     for j in range(1,100):
#         if j % 5 == 0:
#             print("当前爬取页面：第%d页"%j)
#         get_data2(type_list1[i],url_list1[i].format(str(j)))

