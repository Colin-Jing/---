# 这段代码在本地运行
from bs4 import BeautifulSoup
from selenium import webdriver #从selenium库中调用webdriver模块
import time
import openpyxl
from selenium.webdriver.chrome.options import Options
import re




def get_data(kind,url):
    chrome_options = Options()
    # 设置chrome浏览器无界面模式
    chrome_options.add_argument('user-agent="Mozilla/5.0 (iPod; U; CPU iPhone OS 2_1 \
                                like Mac OS X; ja-jp) AppleWebKit/525.18.1 (KHTML, like \
                                Gecko) Version/3.1.1 Mobile/5F137 Safari/525.20"')
    driver = webdriver.Chrome(options=chrome_options) # 设置引擎为Chrome，从本地打开一个Chrome浏览器
    driver.get(url) # 访问页面
    time.sleep(0.5)
    for i in range(50):
        driver.execute_script('window.scrollTo(0, document.body.scrollHeight)')
        if i % 10 == 0:
            print(i)
            try:
                print(driver.find_element_by_link_text('点击返回腾讯网'))
                break
            except:
                print('暂未找到load-more按钮')
        time.sleep(1.5)
    HTML = driver.page_source
    time.sleep(1) # 等待3秒
    bs = BeautifulSoup(HTML,'html.parser')
    label = bs.find_all('img',class_="Monograph")  # 解析网页并提取第一个<label>标签
    try:
        wb = openpyxl.load_workbook(r"./腾讯网/{}.xlsx".format(kind))
        print("./腾讯网/{}.xlsx".format(kind) + "被打开")
    except:
        print("./腾讯网/{}.xlsx不存在！！！将自动创建".format(kind))
        wb = openpyxl.Workbook()
    finally:
        ws = wb.active
    for i in range(len(label)):
        if i % 10 == 0:
            print('{}'.format(i)+'\t'+label[i]['alt']) # 打印label的文本
        ws.append([label[i]['alt'], kind])
    wb.save(r'./腾讯网/{}.xlsx'.format(kind))
    print('./腾讯网/{}.xlsx'.format(kind) + '保存成功！！！')
    wb.close()
    driver.close() # 关闭浏览器

def get_data1(kind,url):
    chrome_options = Options()
    # 设置chrome浏览器无界面模式
    chrome_options.add_argument('--headless')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('log-level=3')
    chrome_options.add_argument('user-agent="Mozilla/5.0 (iPod; U; CPU iPhone OS 2_1 \
                                like Mac OS X; ja-jp) AppleWebKit/525.18.1 (KHTML, like \
                                Gecko) Version/3.1.1 Mobile/5F137 Safari/525.20"')
    driver = webdriver.Chrome(options=chrome_options) # 设置引擎为Chrome，从本地打开一个Chrome浏览器
    driver.get(url) # 访问页面
    for i in range(1500):
        driver.execute_script('window.scrollTo(0, document.body.scrollHeight)')
        if i % 10 == 0:
            print(i)
        time.sleep(0.3)
    HTML = driver.page_source
    time.sleep(1) # 等待3秒
    bs = BeautifulSoup(HTML,'html.parser')
    label = bs.find_all('dd',class_="desc")  # 解析网页并提取第一个<label>标签
    try:
        wb = openpyxl.load_workbook(r"./腾讯网/{}.xlsx".format(kind))
        print("./腾讯网/{}.xlsx".format(kind) + "被打开")
    except:
        print("./腾讯网/{}.xlsx不存在！！！将自动创建".format(kind))
        wb = openpyxl.Workbook()
    finally:
        ws = wb.active
    for i in range(len(label)):
        if i % 10 == 0:
            print('{}'.format(i)+'\t'+label[i].find('a').text) # 打印label的文本
        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
        text = ILLEGAL_CHARACTERS_RE.sub(r'', label[i].find('a').text)
        ws.append([text, kind])
    wb.save(r'./腾讯网/{}.xlsx'.format(kind))
    print('./腾讯网/{}.xlsx'.format(kind) + '保存成功！！！')
    wb.close()
    driver.close() # 关闭浏览器


type_list = ['财经','教育','科技','军事','游戏','娱乐','其他']
url_list = [
    'https://new.qq.com/ch/finance/',
    'https://new.qq.com/ch/edu/',
    'https://new.qq.com/ch/tech/',
    'https://new.qq.com/ch/milite/',
    'https://new.qq.com/ch/games/',
    'https://new.qq.com/ch/ent/',
    'https://new.qq.com/ch/life/'
]
'''
for i in range(len(type_list)):
    print("-*"*20)
    print('当前爬取内容：%s'%type_list[6])
    get_data(type_list[6],url_list[6])
'''

type_list = ['汽车']
url_list = [
    'https://tags.xcar.com.cn/1000502/1/1'
]
# for i in range(len(type_list)):
#     print("-*"*20)
#     print('当前爬取内容：%s'%type_list[i])
#     get_data1(type_list[i], url_list[i])


def get_data2(kind,url):
    chrome_options = Options()
    # 设置chrome浏览器无界面模式
    chrome_options.add_argument('user-agent="Mozilla/5.0 (iPod; U; CPU iPhone OS 2_1 \
                                like Mac OS X; ja-jp) AppleWebKit/525.18.1 (KHTML, like \
                                Gecko) Version/3.1.1 Mobile/5F137 Safari/525.20"')
    driver = webdriver.Chrome(options=chrome_options) # 设置引擎为Chrome，从本地打开一个Chrome浏览器
    driver.get(url) # 访问页面
    driver.execute_script('window.scrollTo(0, document.body.scrollHeight)')
    time.sleep(0.3)
    driver.execute_script('window.scrollTo(0, document.body.scrollHeight)')
    HTML = driver.page_source
    bs = BeautifulSoup(HTML,'html.parser')
    label = bs.find_all('li')  # 解析网页并提取第一个<label>标签
    try:
        wb = openpyxl.load_workbook(r"./腾讯网/{}.xlsx".format(kind))
        print("./腾讯网/{}.xlsx".format(kind) + "被打开")
    except:
        print("./腾讯网/{}.xlsx不存在！！！将自动创建".format(kind))
        wb = openpyxl.Workbook()
    finally:
        ws = wb.active
    for i in range(len(label)):
        if i % 10 == 0:
            print('{}'.format(i)+'\t'+label[i].find('a',target="_blank").text) # 打印label的文本
        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
        text = ILLEGAL_CHARACTERS_RE.sub(r'', label[i].find('a',target="_blank").text)
        ws.append([text, kind])
    wb.save(r'./腾讯网/{}.xlsx'.format(kind))
    print('./腾讯网/{}.xlsx'.format(kind) + '保存成功！！！')
    wb.close()
    driver.close() # 关闭浏览器

def get_data3(kind,url):
    chrome_options = Options()
    # 设置chrome浏览器无界面模式
    chrome_options.add_argument('user-agent="Mozilla/5.0 (iPod; U; CPU iPhone OS 2_1 \
                                like Mac OS X; ja-jp) AppleWebKit/525.18.1 (KHTML, like \
                                Gecko) Version/3.1.1 Mobile/5F137 Safari/525.20"')
    driver = webdriver.Chrome(options=chrome_options) # 设置引擎为Chrome，从本地打开一个Chrome浏览器
    driver.get(url) # 访问页面
    driver.execute_script('window.scrollTo(0, document.body.scrollHeight)')
    time.sleep(0.3)
    driver.execute_script('window.scrollTo(0, document.body.scrollHeight)')
    HTML = driver.page_source
    bs = BeautifulSoup(HTML,'html.parser')
    label = bs.find_all('a',class_='al')  # 解析网页并提取第一个<label>标签
    try:
        wb = openpyxl.load_workbook(r"./腾讯网/{}.xlsx".format(kind))
        print("./腾讯网/{}.xlsx".format(kind) + "被打开")
    except:
        print("./腾讯网/{}.xlsx不存在！！！将自动创建".format(kind))
        wb = openpyxl.Workbook()
    finally:
        ws = wb.active
    for i in range(len(label)):
        if i % 10 == 0:
            print('{}'.format(i)+'\t'+label[i].text) # 打印label的文本
        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
        text = ILLEGAL_CHARACTERS_RE.sub(r'', label[i].text)
        ws.append([text, kind])
    wb.save(r'./腾讯网/{}.xlsx'.format(kind))
    print('./腾讯网/{}.xlsx'.format(kind) + '保存成功！！！')
    wb.close()
    driver.close() # 关闭浏览器

type_list = ['科技']
url_list = [
    'http://scitech.people.com.cn/GB/1057/index{}.html'
]
'''
for i in range(len(type_list)):
    print("-*"*20)
    print('当前爬取内容：%s'%type_list[0])
    for j in range(1,18):
        if j % 5 == 0:
            print("当前爬取页面：第%d页"%j)
        get_data2(type_list[0],url_list[0].format(j))
    for j in range(18,80):
        if j % 5 == 0:
            print("当前爬取页面：第%d页"%j)
        get_data3(type_list[0],url_list[0].format(j))
'''


def get_data4(website,kind,url):
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.128 Safari/537.36")
    driver = webdriver.Chrome(options=chrome_options) # 设置引擎为Chrome，从本地打开一个Chrome浏览器
    driver.get(url) # 访问页面
    time.sleep(0.5)
    for i in range(200):
        driver.execute_script('window.scrollTo(0, document.body.scrollHeight)')
        time.sleep(0.3)
        try:
            if i > 0:
                button = driver.find_element_by_xpath('/html/body/div[5]/div[4]/div[2]/div[1]/div[2]/div/div[3]/div[1]/table/tbody/tr/td/div/ul/li[12]/a')
                button.click()
            else:
                pass
        except:
            pass
        #driver.refresh()
        if i % 10 == 0:
            print(i)
        time.sleep(1.5)
        HTML = driver.page_source
        bs = BeautifulSoup(HTML,'html.parser')
        label = bs.find_all('li')  # 解析网页并提取第一个<label>标签
        try:
            wb = openpyxl.load_workbook(r"./{0}/{1}.xlsx".format(website,kind))
            print("./{0}/{1}.xlsx".format(website, kind) + "被打开")
        except:
            print("./{0}/{1}.xlsx不存在！！！将自动创建".format(website,kind))
            wb = openpyxl.Workbook()
        finally:
            ws = wb.active
        for i in range(len(label)):
            if label[i].find('a', target="_blank") != None:
                if i % 10 == 0:
                    print('{}'.format(i)+'\t'+label[i].find('a',target="_blank").text) # 打印label的文本
                ws.append([label[i].find('a',target="_blank").text, kind])
        wb.save(r'./{0}/{1}.xlsx'.format(website,kind))
        print('./{0}/{1}.xlsx'.format(website,kind)+'保存成功！！！')
        wb.close()
    driver.close() # 关闭浏览器


type_list = ['房产']
url_list = [
   'http://house.hexun.com/list/'
]

# for i in range(len(type_list)):
#     print("-*"*20)
#     print('当前爬取内容：%s'%type_list[i])
#     get_data4('腾讯网',type_list[i],url_list[i])


def get_data5(kind,url):
    chrome_options = Options()
    # 设置chrome浏览器无界面模式
    chrome_options.add_argument('--headless')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('log-level=3')
    chrome_options.add_argument('user-agent="Mozilla/5.0 (iPod; U; CPU iPhone OS 2_1 \
                                like Mac OS X; ja-jp) AppleWebKit/525.18.1 (KHTML, like \
                                Gecko) Version/3.1.1 Mobile/5F137 Safari/525.20"')
    driver = webdriver.Chrome(options=chrome_options) # 设置引擎为Chrome，从本地打开一个Chrome浏览器
    driver.get(url) # 访问页面
    driver.execute_script('window.scrollTo(0, document.body.scrollHeight)')
    time.sleep(0.2)  # 等待3秒
    HTML = driver.page_source

    bs = BeautifulSoup(HTML,'html.parser')
    label = bs.find_all('li')  # 解析网页并提取第一个<label>标签
    print(label)
    try:
        wb = openpyxl.load_workbook(r"./腾讯网/{}.xlsx".format(kind))
        print("./腾讯网/{}.xlsx".format(kind) + "被打开")
    except:
        print("./腾讯网/{}.xlsx不存在！！！将自动创建".format(kind))
        wb = openpyxl.Workbook()
    finally:
        ws = wb.active
    for i in range(len(label)):
        if label[i].find('a',target="_blank") != None:
            if i % 10 == 0:
                print('{}'.format(i)+'\t'+label[i].find('a',target="_blank").text) # 打印label的文本
            ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
            text = ILLEGAL_CHARACTERS_RE.sub(r'', label[i].find('a',target="_blank").text)
            ws.append([text, kind])
    wb.save(r'./腾讯网/{}.xlsx'.format(kind))
    print('./腾讯网/{}.xlsx'.format(kind) + '保存成功！！！')
    wb.close()
    driver.close() # 关闭浏览器


type_list1 = ['其他']
url_list1 = [
   'https://news.qq.com/newssh/shwx/shehuiwanxiang_{}.htm'
]

for i in range(len(type_list1)):
    print("-*"*20)
    print('当前爬取内容：%s'%type_list1[0])
    for j in range(2,81):
        if j % 5 == 0:
            print("当前爬取页面：第%d页"%j)
        get_data5(type_list1[0],url_list1[0].format(j))